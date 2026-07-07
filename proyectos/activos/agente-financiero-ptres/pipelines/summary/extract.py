from openpyxl import load_workbook


def leer_notas_num_factura_ds(
    ruta: str, num_factura_col: int, codigo_col: int, fila_inicio: int
) -> list[dict]:
    wb = load_workbook(ruta)
    sheet = wb["2026"] if "2026" in wb.sheetnames else wb[wb.sheetnames[0]]
    resultado = []
    for row in range(fila_inicio + 1, sheet.max_row + 1):
        celda = sheet.cell(row=row, column=num_factura_col + 1)
        if celda.comment is None:
            continue
        codigo = sheet.cell(row=row, column=codigo_col + 1).value
        if isinstance(codigo, str) and codigo.strip():
            resultado.append({"codigo": codigo.strip(), "nota": celda.comment.text.strip()})
    return resultado


def leer_provisiones_mes_anterior(wb, hoja: str) -> list[dict]:
    sheet = wb[hoja]
    provisiones = []
    for row in range(13, sheet.max_row + 1):
        cierre = sheet.cell(row=row, column=2).value
        if cierre is None or cierre.strip() != "Provision":
            continue
        provisiones.append({
            "proyecto": sheet.cell(row=row, column=8).value,
            "monto_mxn": sheet.cell(row=row, column=12).value,
            "cc": sheet.cell(row=row, column=5).value,
            "cliente": sheet.cell(row=row, column=6).value,
            "nombre_proyecto": sheet.cell(row=row, column=7).value or "",
            "moneda": sheet.cell(row=row, column=9).value or "MXN",
            "monto_original": sheet.cell(row=row, column=10).value,
            "tc": sheet.cell(row=row, column=11).value or 1,
            "anio": sheet.cell(row=row, column=3).value,
            "periodo": sheet.cell(row=row, column=4).value,
        })
    return provisiones


def leer_tipos_cambio(wb, hoja: str) -> dict:
    sheet = wb[hoja]
    tipos = {}
    for row in range(6, 9):
        moneda = sheet.cell(row=row, column=2).value
        if moneda:
            tipos[moneda.strip().upper()] = sheet.cell(row=row, column=3).value
    return tipos
