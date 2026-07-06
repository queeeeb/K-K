import re

from openpyxl import load_workbook

from pipelines.summary.extract import leer_provisiones_mes_anterior, leer_tipos_cambio
from pipelines.summary.extract_fuentes import (
    extraer_consulting,
    extraer_ds,
    extraer_engineering,
    extraer_facturacion,
)
from pipelines.summary.historial import todos_los_codigos_conocidos
from pipelines.summary.interpret import (
    interpret_consulting,
    interpret_ds,
    interpret_engineering,
    interpret_facturacion,
)


def _cargar_rows(ruta: str, hoja: str | None = None) -> list[list]:
    wb = load_workbook(ruta, data_only=True)
    sheet = wb[hoja] if hoja and hoja in wb.sheetnames else wb[wb.sheetnames[0]]
    return [[cell.value for cell in row] for row in sheet.iter_rows()]


_FORMATO_CODIGO_VALIDO = re.compile(r"^\d{2}gmx\d+\.")

_ABREV_MES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}


def _hoja_desde_mes_iso(mes: str) -> str:
    anio, mes_num = mes.split("-")
    return f"{anio}_{_ABREV_MES[int(mes_num)]}"


def _convertir_a_mxn(provisiones: list[dict], tipos_cambio: dict) -> tuple[list[dict], list[str]]:
    alertas = []
    convertidas = []
    for p in provisiones:
        moneda = (p.get("moneda") or "MXN").upper()
        monto_original = p["monto_mxn"]
        if moneda == "MXN":
            convertidas.append({**p, "moneda": "MXN", "monto_original": monto_original, "tc": 1})
            continue
        tc = tipos_cambio.get(moneda)
        if not tc:
            alertas.append(
                f"Proyecto {p['proyecto']} está en {moneda} pero no hay T/C capturado en el "
                "tablero KPI — se dejó el monto sin convertir, requiere revisión manual."
            )
            convertidas.append({**p, "monto_original": monto_original, "tc": None})
            continue
        convertidas.append({
            **p, "monto_original": monto_original, "tc": tc, "monto_mxn": monto_original * tc,
        })
    return convertidas, alertas


def _separar_sospechosos(provisiones: list[dict]) -> tuple[list[dict], list[str]]:
    validas, alertas = [], []
    for p in provisiones:
        if not _FORMATO_CODIGO_VALIDO.match(p["proyecto"]):
            alertas.append(
                f"Código con formato sospechoso excluido de la extracción automática, "
                f"requiere revisión manual: {p['proyecto']!r}"
            )
        else:
            validas.append(p)
    return validas, alertas


def interpretar_summary(raw_files: dict[str, str], client, mes: str | None = None) -> dict:
    mes_numero = int(mes.split("-")[1]) if mes else None

    wb_base = load_workbook(raw_files["base"], data_only=True, keep_vba=True)
    hojas = wb_base.sheetnames
    hoja_mes_anterior = hojas[-1]

    provisiones_mes_anterior = leer_provisiones_mes_anterior(wb_base, hoja_mes_anterior)
    codigos_conocidos = todos_los_codigos_conocidos(wb_base, hojas)
    tipos_cambio = leer_tipos_cambio(wb_base, hoja_mes_anterior)

    rows_facturacion = _cargar_rows(raw_files["facturacion"], hoja="Detalle")
    estructura_facturacion = interpret_facturacion(rows_facturacion, client)
    facturas_mes = extraer_facturacion(rows_facturacion, estructura_facturacion)

    rows_ds = _cargar_rows(raw_files["ds"], hoja="2026")
    estructura_ds = interpret_ds(rows_ds, client, mes_numero=mes_numero)
    ds_actuales = extraer_ds(rows_ds, estructura_ds)

    rows_engineering = _cargar_rows(raw_files["engineering"], hoja="Hoja1")
    estructura_engineering = interpret_engineering(rows_engineering, client, mes_numero=mes_numero)
    engineering_actuales = extraer_engineering(rows_engineering, estructura_engineering)

    rows_consulting = _cargar_rows(raw_files["consulting"], hoja=mes.replace("-", ".") if mes else None)
    estructura_consulting = interpret_consulting(rows_consulting, client)
    consulting_actuales = extraer_consulting(rows_consulting, estructura_consulting)

    provisiones_convertidas, alertas_conversion = _convertir_a_mxn(
        ds_actuales + engineering_actuales + consulting_actuales, tipos_cambio
    )
    provisiones_actuales, alertas_sospechosos = _separar_sospechosos(provisiones_convertidas)
    alertas = alertas_conversion + alertas_sospechosos

    return {
        "provisiones_mes_anterior": provisiones_mes_anterior,
        "facturas_mes": facturas_mes,
        "provisiones_actuales": provisiones_actuales,
        "codigos_conocidos": codigos_conocidos,
        "alertas": alertas,
        "ruta_base": raw_files["base"],
        "hoja_mes_anterior": hoja_mes_anterior,
        "hoja_mes_nuevo": _hoja_desde_mes_iso(mes) if mes else None,
    }
