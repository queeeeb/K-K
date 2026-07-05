from openpyxl import Workbook

from pipelines.summary.historial import buscar_codigo_en_historial, todos_los_codigos_conocidos


def _wb_con_hojas(hojas: dict[str, list[str]]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, codigos in hojas.items():
        sheet = wb.create_sheet(nombre)
        for offset, codigo in enumerate(codigos, start=13):
            sheet.cell(row=offset, column=8, value=codigo)
    return wb


def test_buscar_codigo_encontrado_en_hoja_mas_reciente():
    wb = _wb_con_hojas({
        "2026_Mar": ["26gmx3000.001"],
        "2026_Abr": ["26gmx3000.001"],
    })

    hoja = buscar_codigo_en_historial(wb, ["2026_Mar", "2026_Abr"], "26gmx3000.001")

    assert hoja == "2026_Abr"


def test_buscar_codigo_no_encontrado_devuelve_none():
    wb = _wb_con_hojas({"2026_Mar": ["26gmx3000.001"]})

    hoja = buscar_codigo_en_historial(wb, ["2026_Mar"], "26gmx9999.999")

    assert hoja is None


def test_buscar_codigo_cancelado_devuelve_ultima_aparicion():
    wb = _wb_con_hojas({
        "2026_Ene": ["26gmx3000.001"],
        "2026_Feb": ["26gmx3000.001"],
        "2026_Mar": [],
    })

    hoja = buscar_codigo_en_historial(wb, ["2026_Ene", "2026_Feb", "2026_Mar"], "26gmx3000.001")

    assert hoja == "2026_Feb"


def test_todos_los_codigos_conocidos_junta_todas_las_hojas():
    wb = _wb_con_hojas({
        "2026_Mar": ["26gmx3000.001"],
        "2026_Abr": ["26gmx7000.002", "26gmx3000.001"],
    })

    codigos = todos_los_codigos_conocidos(wb, ["2026_Mar", "2026_Abr"])

    assert codigos == {"26gmx3000.001", "26gmx7000.002"}


def test_todos_los_codigos_conocidos_ignora_celdas_vacias():
    wb = _wb_con_hojas({"2026_Mar": []})

    codigos = todos_los_codigos_conocidos(wb, ["2026_Mar"])

    assert codigos == set()
