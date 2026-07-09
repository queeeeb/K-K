from pathlib import Path

from openpyxl import load_workbook

from pipelines.summary.write import escribir_hoja_mes

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def _escribir(destino, filas, concentrado=None, mes_actual="Mayo"):
    return escribir_hoja_mes(
        ruta_origen=str(FIXTURES_DIR / "summary_abril.xlsm"),
        ruta_destino=str(destino),
        hoja_mes_anterior="2026_Abr",
        hoja_mes_nuevo="2026_May",
        filas=filas,
        concentrado=concentrado if concentrado is not None else {},
        mes_actual=mes_actual,
    )


def test_write_borra_bloque_tipos_cambio(tmp_path):
    destino = tmp_path / "out.xlsm"
    _escribir(destino, [])
    ws = load_workbook(destino, data_only=False, keep_vba=True)["2026_May"]
    for row in range(6, 9):
        for col in range(2, 5):
            assert ws.cell(row=row, column=col).value is None


def test_escribir_hoja_mes_no_toca_columna_a_del_kpi(tmp_path):
    destino = tmp_path / "summary_mayo.xlsm"
    filas = [
        ["Q-3", "Provision", 2026, "Mayo", 2000, "Cliente Cuatro", "Proyecto Cuatro", "26gmx2000.005",
         "MXN", 3000, 1, 3000, 0, 3000, 0, 0, 3000, "", ""],
    ]

    _escribir(destino, filas)

    wb = load_workbook(destino)
    nueva = wb["2026_May"]

    for row in range(1, 12):
        assert nueva.cell(row=row, column=1).value == f"KPI fila {row}"

    assert nueva.cell(row=12, column=1).value == "Cotizacion"
    assert nueva.cell(row=13, column=8).value == "26gmx2000.005"
    assert nueva.cell(row=14, column=1).value is None


def test_escribir_hoja_mes_actualiza_rango_de_formulas_kpi(tmp_path):
    destino = tmp_path / "summary_mayo.xlsm"
    filas = [
        ["Q-1", "Provision", 2026, "Mayo", 3000, "Cliente Uno", "Proyecto Uno", "26gmx3000.001",
         "MXN", 1000, 1, 1000, 0, 1000, 0, 0, 1000, "", ""],
        ["Q-2", "Cancelar", 2026, "Mayo", 7000, "Cliente Dos", "Proyecto Dos", "26gmx7000.002",
         "MXN", 2000, 1, 2000, "", 2000, "", "", 2000, "", ""],
    ]

    _escribir(destino, filas)

    wb = load_workbook(destino)
    nueva = wb["2026_May"]
    ultima_fila = 12 + len(filas)

    assert nueva.cell(row=2, column=9).value == (
        f'=SUMIFS(L13:L{ultima_fila},E13:E{ultima_fila},3000,B13:B{ultima_fila},"Provision")'
    )
    assert nueva.cell(row=2, column=10).value == (
        f'=SUMIFS(L13:L{ultima_fila},E13:E{ultima_fila},2000,B13:B{ultima_fila},"Provision")'
    )
    assert nueva.cell(row=2, column=11).value == (
        f'=SUMIFS(L13:L{ultima_fila},E13:E{ultima_fila},7000,B13:B{ultima_fila},"Provision")'
    )
    assert nueva.cell(row=4, column=9).value == f'=SUMIF(E13:Q{ultima_fila},3000,Q13:Q{ultima_fila})'
    assert nueva.cell(row=4, column=10).value == f'=SUMIF(E13:Q{ultima_fila},2000,Q13:Q{ultima_fila})'
    assert nueva.cell(row=4, column=11).value == f'=SUMIF(E13:Q{ultima_fila},7000,Q13:Q{ultima_fila})'


def test_escribir_hoja_mes_sin_filas_usa_rango_seguro(tmp_path):
    destino = tmp_path / "summary_mayo.xlsm"

    _escribir(destino, [])

    wb = load_workbook(destino)
    nueva = wb["2026_May"]
    assert nueva.cell(row=2, column=9).value == '=SUMIFS(L13:L13,E13:E13,3000,B13:B13,"Provision")'


def test_escribir_hoja_mes_preserva_hoja_anterior(tmp_path):
    destino = tmp_path / "summary_mayo.xlsm"

    _escribir(destino, [])

    wb = load_workbook(destino)
    abril = wb["2026_Abr"]
    assert abril.cell(row=13, column=8).value == "26gmx3000.001"


def _fila(cierre, cc, periodo, monto, proyecto="26gmx3000.001"):
    return ["", cierre, 2026, periodo, cc, "Cli", "Nom", proyecto, "MXN",
            monto, 1, monto, "", "", "", "", "", "", ""]


def test_write_puebla_kpi_filas_3_y_5_desde_concentrado(tmp_path):
    concentrado = {3000: {"facturado": 6050968.44, "canceladas": 552468.62},
                   2000: {"facturado": 1784778.64, "canceladas": 0},
                   7000: {"facturado": 5871693.18, "canceladas": 298668.65}}
    destino = tmp_path / "out.xlsm"
    _escribir(destino, [_fila("Provision", 3000, "Mayo", 1000)], concentrado=concentrado)
    ws = load_workbook(destino, data_only=False, keep_vba=True)["2026_May"]
    assert ws.cell(row=3, column=9).value == 6050968.44
    assert ws.cell(row=5, column=9).value == 552468.62
    assert ws.cell(row=3, column=11).value == 5871693.18


def test_write_concentrado_ausente_deja_blanco_y_alerta(tmp_path):
    destino = tmp_path / "out.xlsm"
    alertas = _escribir(destino, [_fila("Provision", 3000, "Mayo", 1000)], concentrado={})
    ws = load_workbook(destino, data_only=False, keep_vba=True)["2026_May"]
    assert ws.cell(row=3, column=9).value is None
    assert any("Concentrado" in a for a in alertas)


def test_write_fila_11_suma_provisiones_periodo_anterior_por_unidad(tmp_path):
    filas = [
        _fila("Provision", 3000, "Marzo", 100),
        _fila("Provision", 3000, "Mayo", 500),
        _fila("Cancelar", 3000, "Marzo", 999),
    ]
    destino = tmp_path / "out.xlsm"
    _escribir(destino, filas, mes_actual="Mayo")
    ws = load_workbook(destino, data_only=False, keep_vba=True)["2026_May"]
    assert ws.cell(row=11, column=9).value == 100
