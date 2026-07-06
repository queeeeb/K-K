from openpyxl import Workbook

from pipelines.summary.extract import leer_provisiones_mes_anterior, leer_tipos_cambio

HEADERS = ["Cotizacion", "Cierre", "Año", "Periodo", "CC", "Cliente", "Nombre Proyecto",
           "Proyecto", "Moneda", "Provision", "T/C Provision", "PROVISION MXN", "usd",
           "MXN", "EUR", "CAD", "TOTAL MXN", "Referencia", "Comentarios"]


def _wb_con_filas(filas: list[list]) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "2026_Abr"
    for col, header in enumerate(HEADERS, start=1):
        sheet.cell(row=12, column=col, value=header)
    for offset, fila in enumerate(filas, start=13):
        for col, valor in enumerate(fila, start=1):
            sheet.cell(row=offset, column=col, value=valor)
    return wb


def _fila(cierre, cc, cliente, proyecto, monto_mxn, moneda="MXN", provision=None, tc=1, anio=2026, periodo="Abril"):
    provision = monto_mxn if provision is None else provision
    return ["Q-1", cierre, anio, periodo, cc, cliente, "Nombre X", proyecto, moneda,
            provision, tc, monto_mxn, 0, monto_mxn, 0, 0, monto_mxn, "", ""]


def test_leer_provisiones_incluye_solo_cierre_provision():
    wb = _wb_con_filas([
        _fila("Provision", 3000, "Cliente Uno", "26gmx3000.001", 1000),
        _fila("Cancelar", 7000, "Cliente Dos", "26gmx7000.002", 2000),
    ])

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert len(resultado) == 1
    assert resultado[0]["proyecto"] == "26gmx3000.001"


def test_leer_provisiones_normaliza_espacio_en_cierre():
    wb = _wb_con_filas([
        _fila("Provision ", 3000, "Cliente Uno", "26gmx3000.001", 1000),
    ])

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert len(resultado) == 1


def test_leer_provisiones_mapea_campos_esperados():
    wb = _wb_con_filas([
        _fila("Provision", 3000, "Cliente Uno", "26gmx3000.001", 17950),
    ])

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert resultado[0] == {
        "proyecto": "26gmx3000.001",
        "monto_mxn": 17950,
        "cc": 3000,
        "cliente": "Cliente Uno",
        "nombre_proyecto": "Nombre X",
        "moneda": "MXN",
        "monto_original": 17950,
        "tc": 1,
        "anio": 2026,
        "periodo": "Abril",
    }


def test_leer_provisiones_conserva_anio_y_periodo_de_apertura():
    wb = _wb_con_filas([
        _fila("Provision", 3000, "P3 USA", "24gmx3000.104", 93249, anio=2024, periodo="Octubre"),
    ])

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert resultado[0]["anio"] == 2024
    assert resultado[0]["periodo"] == "Octubre"


def test_leer_provisiones_conserva_moneda_original_y_tc():
    wb = _wb_con_filas([
        _fila("Provision", 3000, "Cliente Uno", "26gmx3000.001", 17950, moneda="USD", provision=1000, tc=17.95),
    ])

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert resultado[0]["moneda"] == "USD"
    assert resultado[0]["monto_original"] == 1000
    assert resultado[0]["tc"] == 17.95
    assert resultado[0]["monto_mxn"] == 17950


def test_leer_tipos_cambio_lee_tablero_kpi():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "2026_Abr"
    sheet.cell(row=6, column=2, value="USD")
    sheet.cell(row=6, column=3, value=17.3213)
    sheet.cell(row=7, column=2, value="EUR")
    sheet.cell(row=7, column=3, value=20.2012)
    sheet.cell(row=8, column=2, value="CAD")
    sheet.cell(row=8, column=3, value=None)

    resultado = leer_tipos_cambio(wb, "2026_Abr")

    assert resultado == {"USD": 17.3213, "EUR": 20.2012, "CAD": None}


def test_leer_provisiones_ignora_filas_vacias():
    wb = _wb_con_filas([
        _fila("Provision", 3000, "Cliente Uno", "26gmx3000.001", 1000),
    ])
    # fila de nota suelta sin estructura, columna B vacía
    wb["2026_Abr"].cell(row=14, column=7, value="Se facturó junto con Mayo")

    resultado = leer_provisiones_mes_anterior(wb, "2026_Abr")

    assert len(resultado) == 1
