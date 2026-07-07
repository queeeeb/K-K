"""Generates synthetic .xlsx/.xlsm fixtures with the real column layout but invented data.
Run with: uv run python pipelines/summary/tests/fixtures/make_fixtures.py
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent


def make_summary_abril() -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "2026_Abr"

    for row in range(1, 12):
        sheet.cell(row=row, column=1, value=f"KPI fila {row}")

    headers = ["Cotizacion", "Cierre", "Año", "Periodo", "CC", "Cliente", "Nombre Proyecto",
               "Proyecto", "Moneda", "Provision", "T/C Provision", "PROVISION MXN", "usd",
               "MXN", "EUR", "CAD", "TOTAL MXN", "Referencia", "Comentarios"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=12, column=col, value=header)

    rows = [
        ["Q-1", "Provision", 2026, "Abril", 3000, "Cliente Uno", "Proyecto Uno",
         "26gmx3000.001", "USD", 1000, 17.95, 17950, 1000, 0, 0, 0, 17950, "", ""],
        ["Q-2", "Provision", 2026, "Abril", 7000, "Cliente Dos", "Proyecto Dos",
         "26gmx7000.002", "MXN", 5000, 1, 5000, 0, 5000, 0, 0, 5000, "", ""],
    ]
    for offset, row_values in enumerate(rows, start=13):
        for col, value in enumerate(row_values, start=1):
            sheet.cell(row=offset, column=col, value=value)

    wb.save(FIXTURES_DIR / "summary_abril.xlsm")


def make_facturacion_mayo() -> None:
    wb = Workbook()
    detalle = wb.active
    detalle.title = "Detalle"
    detalle.append(["Proyecto", "Estado", "Periodo", "Monto"])
    detalle.append(["26gmx3000.001-Cliente Uno- Proyecto Uno", "Pagado", datetime(2026, 4, 30), 1000])
    detalle.append(["26gmx7000.099-Cliente Tres- Proyecto Tres", "Cancelado", datetime(2026, 4, 30), 2000])

    concentrado = wb.create_sheet("Concentrado")
    concentrado["A2"], concentrado["B2"] = "CONSULTING", 1000
    concentrado["A3"], concentrado["B3"] = "CANCELADAS", 100
    concentrado["D2"], concentrado["E2"] = "DIGITAL SERVICES", 2000
    concentrado["D3"], concentrado["E3"] = "CANCELADAS", 0
    concentrado["G2"], concentrado["H2"] = "ENGINEERING", 300
    concentrado["G3"], concentrado["H3"] = "CANCELADAS", 0

    wb.save(FIXTURES_DIR / "facturacion_mayo.xlsx")


def make_provisiones_ds_mayo() -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "2026"
    sheet.cell(row=1, column=1, value="Proyecto")
    sub_headers = ["PROVISION", "NUM.FACTURA", "MONTO", "Diferencia+", "Diferencia-", "Acumulados"]
    for col, header in enumerate(sub_headers, start=2):
        sheet.cell(row=2, column=col, value=f"Mayo {header}")
    sheet.cell(row=1, column=7, value="Cliente")
    sheet.cell(row=1, column=8, value="Nombre")
    sheet.cell(row=3, column=1, value="26gmx7000.002")
    sheet.cell(row=3, column=2, value=5000)
    sheet.cell(row=3, column=7, value="Cliente Dos")
    sheet.cell(row=3, column=8, value="Proyecto Dos DS")

    wb.save(FIXTURES_DIR / "provisiones_ds_mayo.xlsx")


def make_provisiones_engineering_mayo() -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Hoja1"
    sheet.append(["Proyecto", "Jan", "Feb", "Mar", "Apr", "May", "Nombre"])
    sheet.append(["26gmx2000.005-Cliente Cuatro", 0, 0, 0, 0, 3000, "Proyecto Cuatro Eng"])

    wb.save(FIXTURES_DIR / "provisiones_engineering_mayo.xlsx")


def make_overview_consulting_mayo() -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "2026.05"
    sheet.append(["STATUS", "PROJECT", "Consultor", "Total", "$"])
    sheet.append(["PROVISION", "26gmx3000.001\nCliente Uno\nProyecto Uno", "Gerardo", 500, "MXN"])
    sheet.append(["", "", "Total honorarios", 600, None])

    wb.save(FIXTURES_DIR / "overview_consulting_mayo.xlsx")


if __name__ == "__main__":
    make_summary_abril()
    make_facturacion_mayo()
    make_provisiones_ds_mayo()
    make_provisiones_engineering_mayo()
    make_overview_consulting_mayo()
    print("Fixtures written to", FIXTURES_DIR)
