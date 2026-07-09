from openpyxl import Workbook

from pipelines.summary.concentrado import leer_concentrado


def _archivo_concentrado(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Concentrado"
    ws["A2"], ws["B2"] = "CONSULTING", 6050968.44
    ws["A3"], ws["B3"] = "CANCELADAS", 552468.62
    ws["A5"], ws["B5"] = "N.CRÉDITO", 12000
    ws["D2"], ws["E2"] = "DIGITAL SERVICES", 5871693.18
    ws["D3"], ws["E3"] = "CANCELADAS", 0
    ws["D5"], ws["E5"] = "N.CRÉDITO", 0
    ws["G2"], ws["H2"] = "ENGINEERING", 1784778.64
    ws["G3"], ws["H3"] = "CANCELADAS", 0
    ws["G5"], ws["H5"] = "N.CRÉDITO", 0
    ruta = tmp_path / "fact.xlsx"
    wb.save(ruta)
    return str(ruta)


def test_leer_concentrado_mapea_unidades(tmp_path):
    resultado = leer_concentrado(_archivo_concentrado(tmp_path))
    assert resultado[3000] == {"facturado": 6050968.44, "canceladas": 552468.62, "nc": 12000}
    assert resultado[7000] == {"facturado": 5871693.18, "canceladas": 0, "nc": 0}
    assert resultado[2000] == {"facturado": 1784778.64, "canceladas": 0, "nc": 0}


def test_leer_concentrado_sin_hoja_devuelve_vacio(tmp_path):
    wb = Workbook()
    wb.active.title = "Detalle"
    ruta = tmp_path / "sin.xlsx"
    wb.save(ruta)
    assert leer_concentrado(str(ruta)) == {}
