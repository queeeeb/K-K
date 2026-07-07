from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from core import auth
from core.api import _conn, app
from core.registry import clear_registry, register
from pipelines.summary.spec import build_summary_spec

FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("AGENTE_DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.setenv("AGENTE_JWT_SECRET", "secreto-de-prueba-no-real-32bytes!!")
    clear_registry()

    destino = tmp_path / "summary_mayo.xlsm"

    def fake_interpret(raw_files):
        return {
            "ledger_vivo": [
                {"proyecto": "26gmx4000.001", "anio": 2026, "periodo": "Abril", "monto_mxn": 2000,
                 "cc": 4000, "cliente": "Cliente Dos", "nombre_proyecto": "", "moneda": "MXN",
                 "monto_original": 2000, "tc": 1},
                {"proyecto": "26gmx3000.001", "anio": 2026, "periodo": "Abril", "monto_mxn": 1000,
                 "cc": 3000, "cliente": "Cliente Uno", "nombre_proyecto": "", "moneda": "MXN",
                 "monto_original": 1000, "tc": 1},
            ],
            "cierres": [
                {"codigo": "26gmx3000.001", "anio": 2026, "mes": "Abril", "origen": "facturacion"},
            ],
            "provisiones_actuales": [
                {"proyecto": "26gmx2000.005", "monto_mxn": 3000, "cc": 2000, "cliente": "Cliente Cuatro",
                 "moneda": "MXN", "monto_original": 3000, "tc": 1}
            ],
            "concentrado": {},
        }

    spec = build_summary_spec(
        interpret_override=fake_interpret,
        ruta_origen=str(FIXTURES_DIR / "summary_abril.xlsm"),
        ruta_destino=str(destino),
        hoja_mes_anterior="2026_Abr",
        hoja_mes_nuevo="2026_May",
    )
    register(spec)
    conn = _conn()
    auth.crear_usuario(conn, "luis", "clave123")
    token = auth.create_access_token("luis")
    headers = {"Authorization": f"Bearer {token}"}
    yield TestClient(app), destino, headers
    clear_registry()


def test_procesar_confirmar_escribe_archivo(client):
    test_client, destino, headers = client

    # summary requiere 5 slots subidos; el fake_interpret los ignora, así que basta
    # con archivos dummy para pasar la validación de subida.
    dummy = ("x.xlsx", b"PK\x03\x04", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    files = {slot: dummy for slot in ("base", "facturacion", "ds", "engineering", "consulting")}
    procesar = test_client.post(
        "/procesar/summary", data={"mes": "2026_May"}, files=files, headers=headers
    )
    assert procesar.status_code == 200
    resumen = procesar.json()["resumen"]
    assert len(resumen["mantenidas"]) == 1
    assert len(resumen["cerradas"]) == 1
    assert len(resumen["nuevas"]) == 1
    assert "alertas" in resumen

    confirmar = test_client.post("/confirmar/summary", json={"token": procesar.json()["token"]}, headers=headers)
    assert confirmar.status_code == 200
    reporte = confirmar.json()["reporte"]
    assert reporte["mantenidas"] == 1
    assert reporte["cerradas"] == 1
    assert reporte["nuevas"] == 1
    assert "filas_escritas" in reporte

    wb = load_workbook(destino)
    hoja = wb["2026_May"]
    assert hoja.cell(row=13, column=8).value == "26gmx4000.001"
    assert hoja.cell(row=14, column=8).value == "26gmx3000.001"
    assert hoja.cell(row=14, column=2).value == "Cancelar"
    assert hoja.cell(row=15, column=8).value == "26gmx2000.005"
    for row in range(1, 12):
        assert hoja.cell(row=row, column=1).value == f"KPI fila {row}"


def test_nombrar_actualiza_plan_antes_de_confirmar(client):
    test_client, destino, headers = client

    dummy = ("x.xlsx", b"PK\x03\x04", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    files = {slot: dummy for slot in ("base", "facturacion", "ds", "engineering", "consulting")}
    procesar = test_client.post(
        "/procesar/summary", data={"mes": "2026_May"}, files=files, headers=headers
    )
    token = procesar.json()["token"]
    assert procesar.json()["resumen"]["nuevas"][0]["cliente"] == "Cliente Cuatro"

    nombrar = test_client.post(
        "/nombrar/summary",
        json={"token": token, "nombres": {"26gmx2000.005": "Cliente Cuatro Renombrado"}},
        headers=headers,
    )
    assert nombrar.status_code == 200
    assert nombrar.json()["resumen"]["nuevas"][0]["cliente"] == "Cliente Cuatro Renombrado"

    confirmar = test_client.post("/confirmar/summary", json={"token": token}, headers=headers)
    assert confirmar.status_code == 200

    wb = load_workbook(destino)
    hoja = wb["2026_May"]
    assert hoja.cell(row=15, column=6).value == "Cliente Cuatro Renombrado"


def test_confirmar_sin_nombrar_permite_continuar(client):
    test_client, destino, headers = client

    dummy = ("x.xlsx", b"PK\x03\x04", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    files = {slot: dummy for slot in ("base", "facturacion", "ds", "engineering", "consulting")}
    procesar = test_client.post(
        "/procesar/summary", data={"mes": "2026_May"}, files=files, headers=headers
    )
    token = procesar.json()["token"]

    confirmar = test_client.post("/confirmar/summary", json={"token": token}, headers=headers)
    assert confirmar.status_code == 200
