from openpyxl import load_workbook

_COLUMNAS_UNIDAD = {3000: "B", 7000: "E", 2000: "H"}


def leer_concentrado(ruta: str) -> dict:
    wb = load_workbook(ruta, data_only=True)
    if "Concentrado" not in wb.sheetnames:
        return {}
    ws = wb["Concentrado"]
    resultado = {}
    for unidad, col_valor in _COLUMNAS_UNIDAD.items():
        facturado = ws[f"{col_valor}2"].value or 0
        canceladas = ws[f"{col_valor}3"].value or 0
        nc = ws[f"{col_valor}5"].value or 0
        resultado[unidad] = {"facturado": facturado, "canceladas": canceladas, "nc": nc}
    return resultado
