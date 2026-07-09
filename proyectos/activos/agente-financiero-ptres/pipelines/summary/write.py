from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def _duplicate_sheet(wb, origen_titulo: str, nuevo_titulo: str):
    origen = wb[origen_titulo]
    nueva = wb.copy_worksheet(origen)
    nueva.title = nuevo_titulo
    nueva.freeze_panes = f"A{_FILA_ENCABEZADOS + 1}"
    return nueva


_SIN_RELLENO = PatternFill(fill_type=None)
_FILA_ENCABEZADOS = 12


def _limpiar_formato(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row == _FILA_ENCABEZADOS:
                continue
            cell.fill = _SIN_RELLENO
            cell.font = Font(name=cell.font.name, size=cell.font.size)
            if cell.comment is not None:
                cell.comment = None


def _limpiar_bloque_tc(sheet) -> None:
    for row in range(6, 9):
        for col in range(2, 5):
            sheet.cell(row=row, column=col).value = None


def _limpiar_seccion_b(sheet) -> None:
    for row in sheet.iter_rows(min_row=12, max_row=sheet.max_row):
        for cell in row:
            cell.value = None


_COL_KPI_POR_UNIDAD = {3000: 9, 2000: 10, 7000: 11}


def _actualizar_formulas_kpi(sheet, ultima_fila: int, mes_actual: str) -> None:
    for col, cc in ((9, 3000), (10, 2000), (11, 7000)):
        # Fila 2 (Provisiones) suma solo las del mes tratado: el filtro por
        # Periodo excluye las provisiones arrastradas de meses anteriores, que
        # se contabilizan aparte en la fila 11 (Prov. Antiguas por facturar).
        sheet.cell(row=2, column=col, value=(
            f'=SUMIFS(L13:L{ultima_fila},E13:E{ultima_fila},{cc},'
            f'B13:B{ultima_fila},"Provision",D13:D{ultima_fila},"{mes_actual}")'
        ))
        sheet.cell(row=4, column=col, value=(
            f'=SUMIF(E13:Q{ultima_fila},{cc},Q13:Q{ultima_fila})'
        ))


def _poblar_facturacion_kpi(sheet, concentrado: dict) -> list[str]:
    if not concentrado:
        for col in _COL_KPI_POR_UNIDAD.values():
            sheet.cell(row=3, column=col).value = None
            sheet.cell(row=5, column=col).value = None
        return ["Hoja Concentrado ausente en Facturación — filas 3 y 5 del KPI quedaron en blanco."]
    for unidad, col in _COL_KPI_POR_UNIDAD.items():
        datos = concentrado.get(unidad, {})
        sheet.cell(row=3, column=col, value=datos.get("facturado"))
        sheet.cell(row=5, column=col, value=datos.get("canceladas"))
    return []


def _poblar_antiguas_por_facturar(sheet, filas: list[list], mes_actual: str) -> None:
    suma = {3000: 0.0, 2000: 0.0, 7000: 0.0}
    for fila in filas:
        cierre = fila[1].strip() if isinstance(fila[1], str) else ""
        periodo = fila[3].strip() if isinstance(fila[3], str) else ""
        unidad, monto = fila[4], fila[11]
        if cierre.startswith("Provision") and periodo != mes_actual and unidad in suma:
            if isinstance(monto, (int, float)):
                suma[unidad] += monto
    for unidad, col in _COL_KPI_POR_UNIDAD.items():
        sheet.cell(row=11, column=col, value=suma[unidad] or 0)


def escribir_hoja_mes(
    ruta_origen: str,
    ruta_destino: str,
    hoja_mes_anterior: str,
    hoja_mes_nuevo: str,
    filas: list[list],
    concentrado: dict,
    mes_actual: str,
) -> list[str]:
    wb = load_workbook(ruta_origen, keep_vba=ruta_origen.endswith(".xlsm"))
    nueva = _duplicate_sheet(wb, hoja_mes_anterior, hoja_mes_nuevo)
    _limpiar_seccion_b(nueva)
    _limpiar_bloque_tc(nueva)

    encabezados = [
        "Cotizacion", "Cierre", "Año", "Periodo", "CC", "Cliente", "Nombre Proyecto",
        "Proyecto", "Moneda", "Provision", "T/C Provision", "PROVISION MXN", "usd",
        "MXN", "EUR", "CAD", "TOTAL MXN", "Referencia", "Comentarios",
    ]
    for col, header in enumerate(encabezados, start=1):
        nueva.cell(row=12, column=col, value=header)

    for offset, fila in enumerate(filas, start=13):
        for col, valor in enumerate(fila, start=1):
            nueva.cell(row=offset, column=col, value=valor)

    ultima_fila = 12 + len(filas) if filas else 13
    _actualizar_formulas_kpi(nueva, ultima_fila, mes_actual)
    alertas = _poblar_facturacion_kpi(nueva, concentrado)
    _poblar_antiguas_por_facturar(nueva, filas, mes_actual)
    _limpiar_formato(nueva)

    wb.save(ruta_destino)
    return alertas
