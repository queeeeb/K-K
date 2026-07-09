"""Escritura determinista del P&L con openpyxl: hojas CONSOLIDATED y BY SEGMENT,
siguiendo el orden de secciones y el catálogo de labels de la referencia (macro).
No decide nada — solo vuelca el plan. Ver ESPECIFICACION §2 y §5.
"""
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipelines.pl import referencia
from pipelines.pl.calculate import COLUMNAS_SALIDA, SEGMENTOS, safe_pct

_ASSETS = os.path.join(os.path.dirname(__file__), "assets")
LOGO_ORION = os.path.join(_ASSETS, "orion_logo.png")
LOGO_P3 = os.path.join(_ASSETS, "p3_logo.png")

ETIQUETAS_SEGMENTO = ["BACK OFFICE", "CONSUL OP", "ENGINEERING", "DIGITAL SOLUTIONS"]

FUENTE = "Arial"
FMT_MONTO = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
FMT_PCT = '_(* 0.0%_);_(* (0.0%);_(* "-"??_);_(@_)'

_CEROS = {col: 0.0 for col in COLUMNAS_SALIDA}
_FILL_TITULO = PatternFill("solid", fgColor="99CCFF")
_FILL_HEADER = PatternFill("solid", fgColor="800000")

_FONT_BASE = Font(name=FUENTE, size=9)
_FONT_BOLD = Font(name=FUENTE, size=9, bold=True)
_FONT_HEADER = Font(name=FUENTE, size=9, bold=True, color="FFFFFF")
_FONT_TITULO = Font(name=FUENTE, size=12, bold=True)

_LADO = Side(style="thin", color="D9D9D9")
_BORDE = Border(left=_LADO, right=_LADO, top=_LADO, bottom=_LADO)
_DERECHA = Alignment(horizontal="right")
_IZQUIERDA = Alignment(horizontal="left")
_CENTRO = Alignment(horizontal="center")


def _agrupar_por_label(filas) -> dict:
    acumulado: dict = {}
    for fila in filas:
        montos = acumulado.setdefault(fila["label"], {col: 0.0 for col in COLUMNAS_SALIDA})
        for col in COLUMNAS_SALIDA:
            montos[col] += fila["montos"][col]
    return acumulado


def _escribir_fila(sheet, r, label, montos, base, consolidado):
    etiqueta = sheet.cell(row=r, column=1, value=label)
    etiqueta.font = _FONT_BASE
    etiqueta.alignment = _IZQUIERDA

    def par(col_monto, valor, base_col):
        m = sheet.cell(row=r, column=col_monto, value=valor)
        m.number_format = FMT_MONTO
        m.font = _FONT_BASE
        m.alignment = _DERECHA
        p = sheet.cell(row=r, column=col_monto + 1, value=safe_pct(valor, base_col))
        p.number_format = FMT_PCT
        p.font = _FONT_BASE
        p.alignment = _DERECHA

    if consolidado:
        par(2, montos["TOTAL"], base["TOTAL"])
    else:
        for i, seg in enumerate(SEGMENTOS):
            par(2 + i * 2, montos[seg], base[seg])
        par(10, montos["TOTAL"], base["TOTAL"])


def _escribir_hoja(sheet, plan, periodo, consolidado):
    base = plan["base_ingresos"]
    totales = plan["totales"]
    rubros = plan["rubros"]
    alloc = plan["allocations"]
    ncols = 3 if consolidado else 11

    def fila(label, montos):
        nonlocal r
        _escribir_fila(sheet, r, label, montos, base, consolidado)
        r += 1

    def titulo(texto):
        nonlocal r
        c = sheet.cell(row=r, column=1, value=texto)
        c.font = _FONT_BOLD
        c.alignment = _IZQUIERDA
        r += 1

    def barra(label, montos, fill, blanco=False):
        nonlocal r
        _escribir_fila(sheet, r, label, montos, base, consolidado)
        fuente = _FONT_HEADER if blanco else _FONT_BOLD
        for col in range(1, ncols + 1):
            celda = sheet.cell(row=r, column=col)
            celda.fill = fill
            celda.font = fuente
        r += 1

    r = 1
    tit = sheet.cell(row=r, column=1, value="P-TRES GROUP, S.A.P.I. DE C.V.")
    tit.font = _FONT_TITULO
    tit.alignment = _CENTRO
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    r += 1
    sufijo = " — " if consolidado else " by Segment — "
    sub = sheet.cell(row=r, column=1, value="Profit and Loss Statement" + sufijo + str(periodo))
    sub.font = Font(name=FUENTE, size=10, bold=True)
    sub.alignment = _CENTRO
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    r += 3

    _agregar_logos(sheet)
    header_row = r
    sheet.cell(row=r, column=1, value="DESCRIPTION")
    if consolidado:
        sheet.cell(row=r, column=2, value="TOTAL")
        sheet.cell(row=r, column=3, value="%")
    else:
        for i, etq in enumerate(ETIQUETAS_SEGMENTO):
            sheet.cell(row=r, column=2 + i * 2, value=etq)
            sheet.cell(row=r, column=3 + i * 2, value="%")
        sheet.cell(row=r, column=10, value="TOTAL")
        sheet.cell(row=r, column=11, value="%")
    for col in range(1, ncols + 1):
        c = sheet.cell(row=r, column=col)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _CENTRO
    r += 1

    incomes = rubros["INCOMES"]
    ns = _agrupar_por_label([f for f in incomes if f["grupo"] == referencia.GRUPO_VENTAS_NACIONALES])
    accrued = _agrupar_por_label([f for f in incomes if f["numero"][:8] in ("4110-004", "4110-005")])
    internacionales = [f for f in incomes if f["numero"][:8] == "4110-002" and f["montos"]["TOTAL"] != 0]

    titulo("Incomes")
    titulo("  NATIONAL SALES")
    for cliente in referencia.CATALOGO_NATIONAL_SALES:
        fila("    " + cliente, ns.get(cliente, _CEROS))
    for f in internacionales:
        fila(f["label"], f["montos"])
    for label in referencia.CATALOGO_ACCRUED_REVENUE:
        fila(label, accrued.get(label, _CEROS))
    barra("Incomes Total", totales["INCOMES"], _FILL_TITULO)

    exp = _agrupar_por_label(rubros["EXPENSES"])
    titulo("Expenses")
    for label in referencia.CATALOGO_EXPENSES:
        fila(label, exp.get(label, _CEROS))
    barra("Expenses Total", totales["EXPENSES"], _FILL_TITULO)

    barra("OPERATING PROFIT (OR LOSS) BEFORE ALLOCATIONS", totales["OPERATING_PROFIT"], _FILL_TITULO)

    if not consolidado:
        titulo("ALLOCATION BO & GMBH SERVICES")
        fila("P3 GLOBAL GMBH SERVICES", alloc["gmbh"])
        fila("  ALLOCATION OF BACK OFFICE - EXPENSES", alloc["back_office"])
        barra("OPERATING PROFIT (OR LOSS) - After Allocation",
              totales["OPERATING_AFTER_ALLOCATION"], _FILL_HEADER, blanco=True)

    for titulo_sec, rubro, total_label, catalogo in [
        ("Other Incomes", "OTHER_INCOMES", "Other Incomes Total", referencia.CATALOGO_OTHER_INCOMES),
        ("Other Expenses", "OTHER_EXPENSES", "Other Expenses Total", referencia.CATALOGO_OTHER_EXPENSES),
        ("Accrued Taxes", "ACCRUED_TAXES", "Accrued Taxes Total", referencia.CATALOGO_ACCRUED_TAXES),
    ]:
        agrupado = _agrupar_por_label(rubros[rubro])
        titulo(titulo_sec)
        for label in catalogo:
            fila(label, agrupado.get(label, _CEROS))
        barra(total_label, totales[rubro], _FILL_TITULO)

    barra("NET PROFIT (OR LOSS)", totales["NET_PROFIT"], _FILL_HEADER, blanco=True)

    _aplicar_bordes(sheet, header_row, r - 1, ncols)
    _ajustar_columnas(sheet, consolidado)
    sheet.freeze_panes = "A" + str(header_row + 1)


def _agregar_logos(sheet) -> None:
    sheet.row_dimensions[1].height = 16
    sheet.row_dimensions[2].height = 15
    sheet.row_dimensions[3].height = 34
    sheet.row_dimensions[4].height = 34

    orion = Image(LOGO_ORION)
    orion.width, orion.height = 108, 32
    orion.anchor = "A1"
    sheet.add_image(orion)

    p3 = Image(LOGO_P3)
    p3.width, p3.height = 60, 65
    p3.anchor = "A3"
    sheet.add_image(p3)


def _aplicar_bordes(sheet, fila_ini, fila_fin, ncols) -> None:
    for row in range(fila_ini, fila_fin + 1):
        for col in range(1, ncols + 1):
            sheet.cell(row=row, column=col).border = _BORDE


def _ajustar_columnas(sheet, consolidado) -> None:
    max_len = max(
        (len(str(sheet.cell(row=r, column=1).value)) for r in range(1, sheet.max_row + 1)
         if sheet.cell(row=r, column=1).value is not None),
        default=0,
    )
    sheet.column_dimensions["A"].width = min(max_len + 2, 46)
    ncols = 3 if consolidado else 11
    for col in range(2, ncols + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14 if col % 2 == 0 else 8


def escribir_pl(ruta_destino: str, plan: dict, periodo: str) -> None:
    wb = Workbook()
    consolidado = wb.active
    consolidado.title = "CONSOLIDATED"
    _escribir_hoja(consolidado, plan, periodo, consolidado=True)

    por_segmento = wb.create_sheet("BY SEGMENT")
    _escribir_hoja(por_segmento, plan, periodo, consolidado=False)

    wb.save(ruta_destino)
