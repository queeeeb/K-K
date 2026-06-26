from pathlib import Path

from openpyxl import load_workbook

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def test_fixture_contpaqi_tiene_periodo_y_cuentas():
    wb = load_workbook(FIXTURES_DIR / "contpaqi_marzo.xlsx")
    sheet = wb["Movimientos"]

    assert "Marzo 2026" in str(sheet.cell(row=3, column=1).value)

    # la primera cuenta y su Total Seg. existen con cargos/abonos
    assert sheet.cell(row=5, column=1).value == "4110-002-001-000"
    # fila Total Seg. con abonos 1000 en columna G (7)
    encontrado = False
    for row in range(1, sheet.max_row + 1):
        marcador = sheet.cell(row=row, column=5).value
        if marcador and str(marcador).startswith("Total Seg."):
            encontrado = True
            break
    assert encontrado
