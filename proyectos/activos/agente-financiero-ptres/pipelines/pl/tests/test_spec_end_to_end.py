import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from core import auth
from core.api import _conn, app
from core.registry import clear_registry, register
from pipelines.pl.spec import build_pl_spec


def _buscar_fila(sheet, label):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"No se encontró la fila '{label}'")


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("AGENTE_DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.setenv("AGENTE_JWT_SECRET", "secreto-de-prueba-no-real-32bytes!!")
    clear_registry()

    destino = tmp_path / "PL_marzo.xlsx"

    def fake_interpret(raw_files):
        return {
            "cuentas": [
                {"numero": "4110-002-001-000", "label": "  FORD", "rubro": "INCOMES",
                 "segmentos": {"ING": {"cargos": 0, "abonos": 1000}}},
                {"numero": "6100-001-001-000", "label": "  GENERAL DEP", "rubro": "EXPENSES",
                 "segmentos": {"ING": {"cargos": 400, "abonos": 0}}},
                {"numero": "4210-001-000-000", "label": "  E.R. FLUCTUATION PROFIT", "rubro": "OTHER_INCOMES",
                 "segmentos": {"ING": {"cargos": 0, "abonos": 50}}},
                {"numero": "6100-009-001-000", "label": "  E.R. FLUCTUATION LOSS", "rubro": "OTHER_EXPENSES",
                 "segmentos": {"ING": {"cargos": 20, "abonos": 0}}},
                {"numero": "8000-001-000-000", "label": "  INCOME TAX OF THE YEAR", "rubro": "ACCRUED_TAXES",
                 "segmentos": {"ING": {"cargos": 10, "abonos": 0}}},
            ]
        }

    spec = build_pl_spec(
        interpret_override=fake_interpret,
        ruta_destino=str(destino),
        periodo="Marzo 2026",
    )
    register(spec)
    conn = _conn()
    auth.crear_usuario(conn, "luis", "clave123")
    token = auth.create_access_token("luis")
    headers = {"Authorization": f"Bearer {token}"}
    yield TestClient(app), destino, headers
    clear_registry()


def test_procesar_confirmar_escribe_pl(client):
    test_client, destino, headers = client

    # pl requiere el slot 'movimientos' subido; el fake_interpret lo ignora, así que
    # basta con un archivo dummy para pasar la validación de subida.
    dummy = ("x.xlsx", b"PK\x03\x04", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    procesar = test_client.post(
        "/procesar/pl", data={"mes": "2026_Mar"}, files={"movimientos": dummy}, headers=headers
    )
    assert procesar.status_code == 200
    resumen = procesar.json()["resumen"]
    assert resumen["incomes"] == 1000
    assert resumen["net_profit"] == 620

    confirmar = test_client.post("/confirmar/pl", json={"token": procesar.json()["token"]}, headers=headers)
    assert confirmar.status_code == 200

    wb = load_workbook(destino)
    cons = wb["CONSOLIDATED"]
    r = _buscar_fila(cons, "NET PROFIT (OR LOSS)")
    assert cons.cell(row=r, column=2).value == 620
