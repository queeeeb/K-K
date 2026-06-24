from openpyxl import load_workbook

from pipelines.pl.calculate import calcular_pl
from pipelines.pl.write import escribir_pl


def _buscar_fila(sheet, label):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"No se encontró la fila '{label}'")


def _plan_demo():
    cuentas = [
        {"numero": "4110-002-001-000", "label": "  FORD", "rubro": "INCOMES",
         "segmentos": {"ING": {"cargos": 0, "abonos": 1000}}},
        {"numero": "6100-001-001-000", "label": "  GENERAL DEP", "rubro": "EXPENSES",
         "segmentos": {"ING": {"cargos": 400, "abonos": 0}}},
        {"numero": "4210-001-000-000", "label": "  E.R. FLUCTUATION PROFIT", "rubro": "OTHER_INCOMES",
         "segmentos": {"ING": {"cargos": 0, "abonos": 50}}},
        {"numero": "6100-009-001-000", "label": "  E.R. FLUCTUATION LOSS", "rubro": "OTHER_EXPENSES",
         "segmentos": {"ING": {"cargos": 20, "abonos": 0}}},
        {"numero": "8000-001-000-000", "label": "  INCOME TAX OF THE YEAR", "rubro": "ACCRUED_TAXES",
         "segmentos": {"ING": {"cargos": 10, "abonos": 0}}},
    ]
    return calcular_pl(cuentas)


def test_escribir_pl_genera_dos_hojas(tmp_path):
    destino = tmp_path / "PL_marzo.xlsx"
    escribir_pl(str(destino), _plan_demo(), periodo="Marzo 2026")

    wb = load_workbook(destino)
    assert wb.sheetnames == ["CONSOLIDATED", "BY SEGMENT"]


def test_consolidated_net_profit(tmp_path):
    destino = tmp_path / "PL_marzo.xlsx"
    escribir_pl(str(destino), _plan_demo(), periodo="Marzo 2026")

    wb = load_workbook(destino)
    cons = wb["CONSOLIDATED"]
    r = _buscar_fila(cons, "NET PROFIT (OR LOSS)")
    assert cons.cell(row=r, column=2).value == 620

    r_op = _buscar_fila(cons, "OPERATING PROFIT (OR LOSS) BEFORE ALLOCATIONS")
    assert cons.cell(row=r_op, column=2).value == 600


def test_by_segment_columnas(tmp_path):
    destino = tmp_path / "PL_marzo.xlsx"
    escribir_pl(str(destino), _plan_demo(), periodo="Marzo 2026")

    wb = load_workbook(destino)
    seg = wb["BY SEGMENT"]
    # ING es el tercer segmento -> monto en columna 6 (2 + 2*2)
    r = _buscar_fila(seg, "Incomes Total")
    assert seg.cell(row=r, column=6).value == 1000
    # columna TOTAL = 10
    assert seg.cell(row=r, column=10).value == 1000
