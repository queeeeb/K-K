"""Genera un export sintético de Contpaqi con el layout jerárquico real pero datos
inventados (ningún dato real de P-TRES GROUP). Mismo enfoque que los fixtures del Summary.
Run: uv run python pipelines/pl/tests/fixtures/make_fixtures.py
"""
from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent


def make_contpaqi_marzo() -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Movimientos"

    sheet.cell(row=1, column=1, value="P-TRES GROUP, S.A.P.I. DE C.V.")
    sheet.cell(row=2, column=1, value="Movimientos Auxiliares por Segmento de Negocio")
    sheet.cell(row=3, column=1, value="Del 1 al 31 de Marzo 2026")

    estado = {"r": 5}

    def cuenta(numero, nombre):
        r = estado["r"]
        sheet.cell(row=r, column=1, value=numero)
        sheet.cell(row=r, column=2, value=nombre)
        estado["r"] += 1

    def segmento(codigo, nombre):
        r = estado["r"]
        sheet.cell(row=r, column=1, value=f"Segmento: {codigo} {nombre}")
        estado["r"] += 1

    def total_seg(nombre, cargos, abonos):
        r = estado["r"]
        sheet.cell(row=r, column=5, value=f"Total Seg. {nombre}")
        sheet.cell(row=r, column=6, value=cargos)
        sheet.cell(row=r, column=7, value=abonos)
        estado["r"] += 1

    # Ingreso (cliente internacional directo): abonos - cargos = 1000
    cuenta("4110-002-001-000", "FORD MOTOR COMPANY")
    segmento("2000", "ING")
    total_seg("ENGINEERING", 0, 1000)

    # Gasto sueldos: cargos - abonos = 400
    cuenta("6100-001-001-000", "SUELDOS Y SALARIOS")
    segmento("2000", "ING")
    total_seg("ENGINEERING", 400, 0)

    # Other income: abonos - cargos = 50
    cuenta("4210-001-000-000", "UTILIDAD CAMBIARIA")
    segmento("2000", "ING")
    total_seg("ENGINEERING", 0, 50)

    # Other expense: cargos - abonos = 20
    cuenta("6100-009-001-000", "PERDIDA CAMBIARIA")
    segmento("2000", "ING")
    total_seg("ENGINEERING", 20, 0)

    # Accrued tax: cargos - abonos = 10
    cuenta("8000-001-000-000", "I.S.R. DEL EJERCICIO")
    segmento("2000", "ING")
    total_seg("ENGINEERING", 10, 0)

    wb.save(FIXTURES_DIR / "contpaqi_marzo.xlsx")


if __name__ == "__main__":
    make_contpaqi_marzo()
    print("Fixtures written to", FIXTURES_DIR)
